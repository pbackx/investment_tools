import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def retirement_savings_planner(current_age, retirement_age, target_amount, current_savings, annual_return, inflation_rate, contribution_frequency="monthly"):
    years_to_retirement = retirement_age - current_age

    # Map contribution frequency to periods
    freq_map = {"daily": 365, "weekly": 52, "bi-weekly": 26, "monthly": 12, "quarterly": 4, "annually": 1}
    periods_per_year = freq_map.get(contribution_frequency.lower(), 12)
    
    # Inflation-adjusted annual return rate
    annual_rate = annual_return / 100
    inflation_adjusted_rate = ((1 + annual_rate) / (1 + inflation_rate / 100)) - 1 if inflation_rate > 0 else annual_rate

    # Calculate contributions per period
    adjusted_target = target_amount / ((1 + inflation_rate / 100) ** years_to_retirement)
    total_periods = years_to_retirement * periods_per_year
    periodic_rate = inflation_adjusted_rate / periods_per_year
    periodic_contribution = (adjusted_target - current_savings * ((1 + periodic_rate) ** total_periods)) / (
        ((1 + periodic_rate) ** total_periods - 1) / periodic_rate
    )

    year_summary = []
    period_details = []
    balance = current_savings

    for year in range(1, years_to_retirement + 1):
        year_start_balance = balance
        total_contributions = 0
        total_interest_earned = 0

        for period in range(periods_per_year):
            interest = balance * periodic_rate
            balance += interest + periodic_contribution
            total_contributions += periodic_contribution
            total_interest_earned += interest

            period_details.append({
                "Year": current_age + year,
                "Period": (year - 1) * periods_per_year + period + 1,
                "Start Balance": balance - interest - periodic_contribution,
                "Contribution": periodic_contribution,
                "Interest Earned": interest,
                "End Balance": balance
            })

        year_summary.append({
            "Year": current_age + year,
            "Start Balance": year_start_balance,
            "Total Contributions": total_contributions,
            "Interest Earned": total_interest_earned,
            "End Balance": balance,
        })

    df_year_summary = pd.DataFrame(year_summary)
    df_period_details = pd.DataFrame(period_details)
    return df_year_summary, df_period_details, periodic_contribution

def plot_retirement_savings(df, file_name):
    # Plot savings progress
    plt.figure(figsize=(12, 7))
    plt.plot(df["Year"], df["End Balance"], label="Total Balance", color="green")
    plt.fill_between(df["Year"], df["Start Balance"], df["End Balance"], alpha=0.2, label="Savings Growth")
    plt.title("Retirement Savings Growth Over Time")
    plt.xlabel("Year")
    plt.ylabel("Balance ($)")
    plt.legend(loc="upper left")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)
    plt.close()

def auto_adjust_column_width(file_name):
    workbook = load_workbook(file_name)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    workbook.save(file_name)

def embed_chart_in_excel(file_name, image_file):
    workbook = load_workbook(file_name)
    graph_sheet_name = "Graph"
    if graph_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(graph_sheet_name)
    chart_sheet = workbook[graph_sheet_name]

    img = Image(image_file)
    img.anchor = "A1"
    chart_sheet.add_image(img)

    workbook.save(file_name)

def export_to_excel(df_year_summary, df_period_details, file_name):
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df_year_summary.to_excel(writer, index=False, sheet_name="Yearly Summary")
        df_period_details.to_excel(writer, index=False, sheet_name="Detailed Breakdown")
    workbook = load_workbook(file_name)
    
    # Apply formatting to the Yearly Summary sheet
    sheet = workbook["Yearly Summary"]
    dollar_columns = ["Start Balance", "Total Contributions", "Interest Earned", "End Balance"]
    for col_name in dollar_columns:
        if col_name in df_year_summary.columns:
            col_letter = sheet.cell(row=1, column=df_year_summary.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    # Apply formatting to the Detailed Breakdown sheet
    sheet = workbook["Detailed Breakdown"]
    dollar_columns = ["Start Balance", "Contribution", "Interest Earned", "End Balance"]
    for col_name in dollar_columns:
        if col_name in df_period_details.columns:
            col_letter = sheet.cell(row=1, column=df_period_details.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    workbook.save(file_name)

if __name__ == "__main__":
    while True:
        try:
            current_age = int(input("Enter your current age: "))
            retirement_age = int(input("Enter your desired retirement age: "))
            target_amount = float(input("Enter your target retirement amount: "))
            current_savings = float(input("Enter your current savings (optional, default is 0): ") or 0)
            annual_return = float(input("Enter the expected annual return rate (in %, e.g., 7): "))
            inflation_rate = float(input("Enter the expected annual inflation rate (optional, default is 0): ") or 0)
            contribution_frequency = input("Enter the contribution frequency ('daily', 'weekly', 'bi-weekly', 'monthly', 'quarterly', or 'annually'): ").lower()
            file_name = input("Enter the base name for the output files (e.g., 'retirement_savings'): ")
            break
        except Exception as e:
            print(f"Error: {e}. Please try again.")

    image_file = f"{file_name}.png"
    excel_file = f"{file_name}.xlsx"

    df_year_summary, df_period_details, periodic_contribution = retirement_savings_planner(
        current_age, retirement_age, target_amount, current_savings, annual_return, inflation_rate, contribution_frequency
    )
    plot_retirement_savings(df_year_summary, image_file)
    export_to_excel(df_year_summary, df_period_details, excel_file)
    embed_chart_in_excel(excel_file, image_file)
    auto_adjust_column_width(excel_file)

    frequency_label = contribution_frequency.capitalize()
    print(f"Retirement savings details saved to {excel_file} with a progress graph embedded.")
    print(f"Required {frequency_label} Contribution: ${periodic_contribution:,.2f}")
