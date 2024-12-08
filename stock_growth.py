import math
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def stock_growth_calculator(initial_investment, annual_rate, contribution, frequency, duration, is_duration_in_years, dividend_yield=0, reinvest_dividends=True):
    # Map contribution frequencies to periods
    freq_map = {"daily": 365, "weekly": 52, "bi-weekly": 26, "monthly": 12, "quarterly": 4, "annually": 1}
    periods_per_year = freq_map.get(frequency, 12)
    
    # Convert duration to periods
    total_periods = duration * periods_per_year if is_duration_in_years else duration

    # Periodic rate of return
    periodic_rate = (annual_rate / 100) / periods_per_year
    dividend_rate = (dividend_yield / 100) / periods_per_year

    results = []
    balance = initial_investment
    total_contributions = initial_investment
    total_growth = 0
    total_dividends = 0

    for period in range(1, total_periods + 1):
        # Apply growth from market returns
        growth = balance * periodic_rate
        balance += growth
        total_growth += growth

        # Apply dividends (reinvested or not)
        dividends = balance * dividend_rate
        if reinvest_dividends:
            balance += dividends
        total_dividends += dividends

        # Add periodic contribution
        balance += contribution
        total_contributions += contribution

        # Save results at each period
        results.append({
            "Period": period,
            "Year": math.ceil(period / periods_per_year),
            "Total Contributions": total_contributions,
            "Dividends Earned (This Period)": dividends,
            "Growth (This Period)": growth,
            "Total Dividends": total_dividends,
            "Total Growth": total_growth,
            "Balance": balance
        })

    # Convert results to DataFrame
    df = pd.DataFrame(results)

    return df

def plot_stock_growth(df, file_name):
    # Convert monetary columns to numeric for plotting
    df["Balance"] = pd.to_numeric(df["Balance"], errors='coerce')
    df["Total Contributions"] = pd.to_numeric(df["Total Contributions"], errors='coerce')
    df["Total Dividends"] = pd.to_numeric(df["Total Dividends"], errors='coerce')

    # Plot the growth over time
    plt.figure(figsize=(12, 7))
    plt.plot(df["Period"], df["Balance"], label="Total Balance", color="blue")
    plt.plot(df["Period"], df["Total Contributions"], label="Total Contributions", linestyle="--", color="orange")
    plt.plot(df["Period"], df["Total Dividends"], label="Total Dividends", linestyle="-.", color="green")
    plt.title("Stock Investment Growth Over Time")
    plt.xlabel("Period")
    plt.ylabel("Balance ($)")
    plt.legend(loc="upper left")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)  # Save the graph to a file
    plt.close()  # Close the plot to avoid pausing the script

def auto_adjust_column_width(file_name):
    # Load workbook
    workbook = load_workbook(file_name)

    # Iterate through all sheets and adjust column widths
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width

    # Save the workbook with adjusted column widths
    workbook.save(file_name)

def embed_chart_in_excel(file_name, image_file):
    # Load workbook and create a new sheet for the chart
    workbook = load_workbook(file_name)
    chart_sheet_name = "Graph"
    if chart_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(chart_sheet_name)
    chart_sheet = workbook[chart_sheet_name]

    # Add the chart to the new sheet
    img = Image(image_file)
    img.anchor = "A1"  # Position the image
    chart_sheet.add_image(img)

    # Save the workbook
    workbook.save(file_name)

def export_to_excel(df, file_name):
    # Export results to an Excel file
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Stock Growth")
    
    # Load the workbook and access the sheet
    workbook = load_workbook(file_name)
    sheet = workbook["Stock Growth"]

    # Apply dollar formatting to specific columns
    dollar_columns = ["Total Contributions", "Dividends Earned (This Period)", "Growth (This Period)", "Total Dividends", "Total Growth", "Balance"]
    for col_name in dollar_columns:
        if col_name in df.columns:
            col_letter = sheet.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):  # Skip header row
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    # Save the workbook with formatting
    workbook.save(file_name)
    print(f"Data exported to {file_name} with dollar formatting.")

if __name__ == "__main__":
    # User inputs
    initial_investment = float(input("Enter the initial investment amount: "))
    annual_rate = float(input("Enter the annual rate of return (in %): "))
    
    # Validate contribution input
    while True:
        try:
            contribution = float(input("Enter the contribution amount per period: "))
            if contribution < 0:
                raise ValueError("Contribution must be a positive number.")
            break
        except ValueError:
            print("Invalid input. Please enter a valid numeric amount for the contribution (e.g., '50').")

    frequency = input("Enter the contribution frequency (daily, weekly, bi-weekly, monthly, quarterly, annually): ").lower()
    if frequency not in {"daily", "weekly", "bi-weekly", "monthly", "quarterly", "annually"}:
        raise ValueError("Invalid frequency. Choose daily, weekly, bi-weekly, monthly, quarterly, or annually.")
    
    # Duration input
    while True:
        duration_input = input("Enter the investment duration (e.g., '12 months' or '5 years'): ").strip().lower()
        try:
            duration_parts = duration_input.split()
            duration = int(duration_parts[0])
            duration_type = duration_parts[1]
            if duration_type not in {"months", "years"}:
                raise ValueError("Invalid unit. Please specify 'months' or 'years'.")
            is_duration_in_years = duration_type == "years"
            break
        except (IndexError, ValueError):
            print("Invalid input. Please enter a number followed by 'months' or 'years' (e.g., '12 months').")

    dividend_yield = float(input("Enter the dividend yield (in %, optional, default is 0): ") or 0)
    reinvest_dividends = input("Do you want dividends reinvested? (yes or no): ").lower() == "yes"
    base_file_name = input("Enter the base name for the output files (e.g., 'results'): ")

    # Generate file names
    graph_file = f"{base_file_name}.png"
    excel_file = f"{base_file_name}.xlsx"

    # Calculate stock growth
    df = stock_growth_calculator(initial_investment, annual_rate, contribution, frequency, duration, is_duration_in_years, dividend_yield, reinvest_dividends)

    # Plot and export results
    plot_stock_growth(df, graph_file)
    export_to_excel(df, excel_file)

    # Embed the graph in the spreadsheet
    embed_chart_in_excel(excel_file, graph_file)

    # Auto-adjust column widths
    auto_adjust_column_width(excel_file)

