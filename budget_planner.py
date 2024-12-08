import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def calculate_budget(monthly_income, budget_categories, actual_spending):
    # Calculate budget details
    results = []
    total_budget = sum(budget_categories.values())
    total_actual = sum(actual_spending.values())
    remaining_income = monthly_income - total_budget

    for category, budget in budget_categories.items():
        actual = actual_spending.get(category, 0)
        results.append({
            "Category": category,
            "Budgeted Amount": budget,
            "Actual Spending": actual,
            "Difference": actual - budget,
            "Percentage of Income": (budget / monthly_income) * 100
        })

    # Summary row
    results.append({
        "Category": "Total",
        "Budgeted Amount": total_budget,
        "Actual Spending": total_actual,
        "Difference": total_actual - total_budget,
        "Percentage of Income": (total_budget / monthly_income) * 100
    })

    df = pd.DataFrame(results)
    return df, remaining_income

def plot_budget_allocation(df, file_name):
    # Pie chart for budget allocation
    categories = df[df["Category"] != "Total"]
    plt.figure(figsize=(8, 8))
    plt.pie(categories["Budgeted Amount"], labels=categories["Category"], autopct="%1.1f%%", startangle=140)
    plt.title("Budget Allocation")
    plt.tight_layout()
    pie_chart_file = file_name.replace(".xlsx", "_pie.png")
    plt.savefig(pie_chart_file)
    plt.close()

    # Bar chart for budget vs. actual spending
    plt.figure(figsize=(10, 6))
    plt.bar(categories["Category"], categories["Budgeted Amount"], label="Budgeted", alpha=0.7)
    plt.bar(categories["Category"], categories["Actual Spending"], label="Actual", alpha=0.7)
    plt.title("Budget vs. Actual Spending")
    plt.xlabel("Category")
    plt.ylabel("Amount ($)")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    bar_chart_file = file_name.replace(".xlsx", "_bar.png")
    plt.savefig(bar_chart_file)
    plt.close()

    return pie_chart_file, bar_chart_file

def auto_adjust_column_width(file_name):
    workbook = load_workbook(file_name)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    workbook.save(file_name)

def embed_charts_in_excel(file_name, pie_chart_file, bar_chart_file):
    workbook = load_workbook(file_name)

    # Add pie chart to a separate sheet
    pie_chart_sheet_name = "Pie Chart"
    if pie_chart_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(pie_chart_sheet_name)
    pie_chart_sheet = workbook[pie_chart_sheet_name]

    pie_img = Image(pie_chart_file)
    pie_img.anchor = "A1"
    pie_chart_sheet.add_image(pie_img)

    # Add bar chart to a separate sheet
    bar_chart_sheet_name = "Bar Chart"
    if bar_chart_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(bar_chart_sheet_name)
    bar_chart_sheet = workbook[bar_chart_sheet_name]

    bar_img = Image(bar_chart_file)
    bar_img.anchor = "A1"
    bar_chart_sheet.add_image(bar_img)

    workbook.save(file_name)


def export_to_excel(df, file_name):
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Budget Summary")
    workbook = load_workbook(file_name)
    sheet = workbook["Budget Summary"]

    # Apply dollar formatting
    dollar_columns = ["Budgeted Amount", "Actual Spending", "Difference"]
    for col_name in dollar_columns:
        if col_name in df.columns:
            col_letter = sheet.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    percentage_column = "Percentage of Income"
    col_letter = sheet.cell(row=1, column=df.columns.get_loc(percentage_column) + 1).column_letter
    for row in range(2, sheet.max_row + 1):
        cell = sheet[f"{col_letter}{row}"]
        cell.number_format = '0.00"%"'  # Percentage format

    workbook.save(file_name)

if __name__ == "__main__":
    while True:
        try:
            monthly_income = float(input("Enter your monthly income: "))
            print("Enter your budget categories and their amounts (e.g., 'Housing: 1500'). Type 'done' when finished.")
            budget_categories = {}
            while True:
                entry = input("Category and amount: ")
                if entry.lower() == "done":
                    break
                try:
                    category, amount = entry.split(":")
                    budget_categories[category.strip()] = float(amount.strip())
                except ValueError:
                    print("Invalid input. Use the format 'Category: Amount'.")

            print("Enter your actual spending in each category (e.g., 'Housing: 1400'). Type 'done' when finished.")
            actual_spending = {}
            while True:
                entry = input("Category and amount: ")
                if entry.lower() == "done":
                    break
                try:
                    category, amount = entry.split(":")
                    actual_spending[category.strip()] = float(amount.strip())
                except ValueError:
                    print("Invalid input. Use the format 'Category: Amount'.")

            file_name = input("Enter the base name for the output files (e.g., 'budget_report'): ")
            if not file_name.endswith(".xlsx"):
                file_name += ".xlsx"
            break
        except Exception as e:
            print(f"Error: {e}. Please try again.")

    df, remaining_income = calculate_budget(monthly_income, budget_categories, actual_spending)
    pie_chart_file, bar_chart_file = plot_budget_allocation(df, file_name)
    export_to_excel(df, file_name)
    embed_charts_in_excel(file_name, pie_chart_file, bar_chart_file)
    auto_adjust_column_width(file_name)

    print(f"Budget details saved to {file_name} with charts embedded.")
    if remaining_income > 0:
        print(f"Remaining Income: ${remaining_income:,.2f}")
    else:
        print(f"Over Budget by: ${-remaining_income:,.2f}")
