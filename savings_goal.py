import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def calculate_savings_goal(target_amount, current_savings, duration, is_years, return_rate, inflation_rate, contribution_frequency):
    # Convert duration to months or years
    if is_years:
        total_months = duration * 12
    else:
        total_months = duration

    # Map contribution frequency to periods
    freq_map = {"daily": 365, "weekly": 52, "bi-weekly": 26, "monthly": 12, "quarterly": 4, "annually": 1}
    periods_per_year = freq_map.get(contribution_frequency.lower(), 12)
    total_periods = (total_months / 12) * periods_per_year

    # Calculate inflation-adjusted goal
    adjusted_goal = target_amount / ((1 + inflation_rate / 100) ** (total_months / 12))

    # Periodic return rate
    periodic_rate = (1 + return_rate / 100) ** (1 / periods_per_year) - 1

    # Calculate required contribution per period
    contribution_per_period = (adjusted_goal - current_savings * (1 + periodic_rate) ** total_periods) / (
        ((1 + periodic_rate) ** total_periods - 1) / periodic_rate
    )

    results = []
    balance = current_savings

    for period in range(1, int(total_periods) + 1):
        interest = balance * periodic_rate
        balance += interest + contribution_per_period

        results.append({
            "Period": period,
            "Year": period // periods_per_year,
            "Contribution": contribution_per_period,
            "Interest Earned": interest,
            "End Balance": balance,
        })

    df = pd.DataFrame(results)
    return df, contribution_per_period

def plot_savings_goal(df, target_amount, file_name):
    # Plot savings progress
    plt.figure(figsize=(12, 7))
    plt.plot(df["Period"], df["End Balance"], label="Savings Balance", color="green")
    plt.axhline(y=target_amount, color="blue", linestyle="--", label="Savings Goal")
    plt.title("Savings Goal Progress")
    plt.xlabel("Period")
    plt.ylabel("Amount ($)")
    plt.legend(loc="upper left")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)
    plt.close()

def auto_adjust_column_width(file_name):
    workbook = load_workbook(file_name)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    workbook.save(file_name)

def embed_chart_in_excel(file_name, image_file):
    workbook = load_workbook(file_name)
    graph_sheet_name = "Graph"
    if graph_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(graph_sheet_name)
    chart_sheet = workbook[graph_sheet_name]

    img = Image(image_file)
    img.anchor = "A1"
    chart_sheet.add_image(img)

    workbook.save(file_name)

def export_to_excel(df, file_name):
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Savings Goal Progress")
    workbook = load_workbook(file_name)
    sheet = workbook["Savings Goal Progress"]

    # Apply dollar formatting
    dollar_columns = ["Contribution", "Interest Earned", "End Balance"]
    for col_name in dollar_columns:
        if col_name in df.columns:
            col_letter = sheet.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    workbook.save(file_name)

if __name__ == "__main__":
    while True:
        try:
            target_amount = float(input("Enter your savings goal: "))
            current_savings = float(input("Enter your current savings (optional, default is 0): ") or 0)
            duration_input = input("Enter the duration to achieve your goal (e.g., '12 months' or '5 years'): ").lower()
            duration_parts = duration_input.split()
            duration = int(duration_parts[0])
            is_years = "year" in duration_parts[1]
            return_rate = float(input("Enter the expected annual return rate (in %, e.g., 5): "))
            inflation_rate = float(input("Enter the expected annual inflation rate (optional, default is 0): ") or 0)
            contribution_frequency = input("Enter the contribution frequency ('daily', 'weekly', 'bi-weekly', 'monthly', 'quarterly', 'annually'): ").lower()
            file_name = input("Enter the base name for the output files (e.g., 'savings_goal'): ")
            break
        except Exception as e:
            print(f"Error: {e}. Please try again.")

    image_file = f"{file_name}.png"
    excel_file = f"{file_name}.xlsx"

    df, periodic_contribution = calculate_savings_goal(
        target_amount, current_savings, duration, is_years, return_rate, inflation_rate, contribution_frequency
    )
    plot_savings_goal(df, target_amount, image_file)
    export_to_excel(df, excel_file)
    embed_chart_in_excel(excel_file, image_file)
    auto_adjust_column_width(excel_file)

    print(f"Savings goal details saved to {excel_file} with a progress graph embedded.")
    print(f"Required {contribution_frequency.capitalize()} Contribution: ${periodic_contribution:,.2f}")
