import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def loan_vs_savings(expense_amount, current_savings, loan_rate, loan_term_years, return_rate, inflation_rate, savings_term_months, savings_frequency="monthly"):
    # Loan scenario calculations
    loan_term_months = loan_term_years * 12
    monthly_loan_rate = (loan_rate / 100) / 12
    monthly_payment = expense_amount * monthly_loan_rate / (1 - (1 + monthly_loan_rate) ** -loan_term_months)
    total_loan_cost = monthly_payment * loan_term_months
    total_loan_interest = total_loan_cost - expense_amount

    # Savings scenario calculations
    freq_map = {"daily": 365, "weekly": 52, "bi-weekly": 26, "monthly": 12}
    periods_per_year = freq_map.get(savings_frequency.lower(), 12)
    savings_periods = savings_term_months * (periods_per_year / 12)
    inflation_adjusted_goal = expense_amount / ((1 + inflation_rate / 100) ** (savings_term_months / 12))
    periodic_rate = (1 + return_rate / 100) ** (1 / periods_per_year) - 1

    # Calculate required contribution per period
    required_contribution = (inflation_adjusted_goal - current_savings * (1 + periodic_rate) ** savings_periods) / (
        ((1 + periodic_rate) ** savings_periods - 1) / periodic_rate
    )

    savings_data = []
    savings_balance = current_savings

    for period in range(1, int(savings_periods) + 1):
        interest = savings_balance * periodic_rate
        savings_balance += interest + required_contribution

        # Calculate time in months and years
        time_in_months = (period / periods_per_year) * 12
        years = int(time_in_months // 12)
        months = int(time_in_months % 12)

        savings_data.append({
            "Period": period,
            "Time (Years/Months)": f"{years}y {months}m" if years > 0 else f"{months}m",
            "Contribution": required_contribution,
            "Interest Earned": interest,
            "Savings Balance": savings_balance,
        })

    total_savings_interest = savings_balance - current_savings - required_contribution * savings_periods

    return {
        "loan": {
            "Monthly Payment": monthly_payment,
            "Total Interest Paid": total_loan_interest,
            "Total Cost": total_loan_cost,
        },
        "savings": {
            "Required Contribution": required_contribution,
            "Total Interest Earned": total_savings_interest,
            "Final Balance": savings_balance,
        },
        "comparison": {
            "Loan Total Cost": total_loan_cost,
            "Savings Final Balance": savings_balance,
        },
        "savings_data": pd.DataFrame(savings_data),
    }

def plot_comparison(loan_cost, savings_balance, savings_data, file_name):
    # Determine the x-axis label based on savings timeframe
    total_periods = len(savings_data)
    if total_periods > 24:  # If long timeframe, use years
        savings_data["Time (Years)"] = savings_data["Period"] / (12 if total_periods > 12 else 1)
        x_label = "Time (Years)"
        x_data = savings_data["Time (Years)"]
    else:
        x_label = "Time (Months)"
        x_data = savings_data["Period"] / 2  # Example: If bi-weekly, convert to months

    plt.figure(figsize=(12, 7))
    plt.axhline(y=loan_cost, color="red", linestyle="--", label="Loan Total Cost")
    plt.plot(x_data, savings_data["Savings Balance"], label="Savings Balance", color="green")
    plt.title("Loan vs Savings Comparison")
    plt.xlabel(x_label)
    plt.ylabel("Amount ($)")
    plt.legend(loc="upper left")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)
    plt.close()

def export_to_excel(loan_results, savings_results, savings_data, file_name):
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        # Loan details
        loan_df = pd.DataFrame([loan_results])
        loan_df.to_excel(writer, index=False, sheet_name="Loan Details")

        # Savings details
        savings_df = pd.DataFrame([savings_results])
        savings_df.to_excel(writer, index=False, sheet_name="Savings Details")

        # Savings breakdown
        savings_data.to_excel(writer, index=False, sheet_name="Savings Breakdown")

    # Apply formatting
    workbook = load_workbook(file_name)

    # Format Loan Details sheet
    sheet = workbook["Loan Details"]
    for col in ["Monthly Payment", "Total Interest Paid", "Total Cost"]:
        if col in loan_df.columns:
            col_letter = sheet.cell(row=1, column=loan_df.columns.get_loc(col) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    # Format Savings Details sheet
    sheet = workbook["Savings Details"]
    for col in ["Required Contribution", "Total Interest Earned", "Final Balance"]:
        if col in savings_df.columns:
            col_letter = sheet.cell(row=1, column=savings_df.columns.get_loc(col) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    # Format Savings Breakdown sheet
    sheet = workbook["Savings Breakdown"]
    for col in ["Contribution", "Interest Earned", "Savings Balance"]:
        if col in savings_data.columns:
            col_letter = sheet.cell(row=1, column=savings_data.columns.get_loc(col) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    # Format time column (Periods and Time)
    if "Period" in savings_data.columns:
        col_letter = sheet.cell(row=1, column=savings_data.columns.get_loc("Period") + 1).column_letter
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"{col_letter}{row}"]
            cell.number_format = '0'
    if "Time (Years/Months)" in savings_data.columns:
        col_letter = sheet.cell(row=1, column=savings_data.columns.get_loc("Time (Years/Months)") + 1).column_letter
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f"{col_letter}{row}"]
            cell.number_format = '@'  # Text format for the time column

    # Auto-adjust column widths for all sheets
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width

    workbook.save(file_name)


def embed_chart_in_excel(file_name, chart_file):
    workbook = load_workbook(file_name)
    chart_sheet_name = "Comparison Chart"
    if chart_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(chart_sheet_name)
    chart_sheet = workbook[chart_sheet_name]

    img = Image(chart_file)
    img.anchor = "A1"
    chart_sheet.add_image(img)

    workbook.save(file_name)

if __name__ == "__main__":
    try:
        # Gather inputs
        expense_amount = float(input("Enter the expense amount: "))
        current_savings = float(input("Enter your current savings (optional, default is 0): ") or 0)
        loan_rate = float(input("Enter the loan interest rate (in %, e.g., 5): "))
        loan_term_years = int(input("Enter the loan term (in years): "))
        return_rate = float(input("Enter the expected annual return rate on investments (in %, e.g., 5): "))
        inflation_rate = float(input("Enter the expected annual inflation rate (in %, optional, default is 0): ") or 0)
        savings_term_months = int(input("Enter the timeframe for saving the expense amount (in months): "))
        savings_frequency = input("Enter the savings contribution frequency ('daily', 'weekly', 'bi-weekly', 'monthly'): ").lower()
        file_name = input("Enter the base name for the output files (e.g., 'savings_vs_loan'): ")

        # Perform calculations
        results = loan_vs_savings(
            expense_amount,
            current_savings,
            loan_rate,
            loan_term_years,
            return_rate,
            inflation_rate,
            savings_term_months,
            savings_frequency
        )

        # Save results and generate outputs
        chart_file = f"{file_name}.png"
        excel_file = f"{file_name}.xlsx"

        # Generate comparison chart
        plot_comparison(results["loan"]["Total Cost"], results["savings"]["Final Balance"], results["savings_data"], chart_file)

        # Export results to Excel
        export_to_excel(results["loan"], results["savings"], results["savings_data"], excel_file)

        # Embed chart into Excel
        embed_chart_in_excel(excel_file, chart_file)

        print(f"Results saved to {excel_file} with a comparison chart embedded.")
        print(f"Loan Total Cost: ${results['loan']['Total Cost']:,.2f}")
        print(f"Savings Final Balance: ${results['savings']['Final Balance']:,.2f}")
        print(f"Required {savings_frequency.capitalize()} Contribution: ${results['savings']['Required Contribution']:,.2f}")

    except Exception as e:
        print(f"Error: {e}. Please try again.")
