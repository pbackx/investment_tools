import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def auto_loan_calculator(loan_amount, interest_rate, loan_term, down_payment=0, trade_in_value=0, extra_payment=0):
    # Subtract down payment and trade-in value from loan amount
    loan_amount -= (down_payment + trade_in_value)

    # Monthly interest rate and total number of payments
    monthly_rate = (interest_rate / 100) / 12
    total_payments = loan_term * 12

    # Calculate monthly payment (standard formula for fixed loans)
    monthly_payment = loan_amount * monthly_rate / (1 - (1 + monthly_rate) ** -total_payments)

    # Amortization schedule
    balance = loan_amount
    total_interest_paid = 0
    results = []

    for month in range(1, total_payments + 1):
        interest = balance * monthly_rate
        principal = monthly_payment - interest
        total_interest_paid += interest
        balance -= (principal + extra_payment)
        balance = max(balance, 0)  # Prevent negative balances

        results.append({
            "Month": month,
            "Monthly Payment": monthly_payment + extra_payment if balance > 0 else 0,
            "Principal Paid": principal + extra_payment if balance > 0 else 0,
            "Interest Paid": interest if balance > 0 else 0,
            "Total Interest Paid": total_interest_paid,
            "Remaining Balance": balance
        })

        # Stop if the loan is paid off early
        if balance <= 0:
            break

    # Convert to DataFrame
    df = pd.DataFrame(results)
    return df

def plot_loan_amortization(df, file_name):
    # Plot principal vs. interest breakdown
    plt.figure(figsize=(12, 7))
    plt.plot(df["Month"], df["Remaining Balance"], label="Remaining Balance", color="blue")
    plt.plot(df["Month"], df["Total Interest Paid"], label="Total Interest Paid", linestyle="--", color="orange")
    plt.title("Loan Amortization Over Time")
    plt.xlabel("Month")
    plt.ylabel("Amount ($)")
    plt.legend(loc="upper right")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)
    plt.close()

def auto_adjust_column_width(file_name):
    workbook = load_workbook(file_name)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    workbook.save(file_name)

def embed_chart_in_excel(file_name, image_file):
    workbook = load_workbook(file_name)
    graph_sheet_name = "Graph"
    if graph_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(graph_sheet_name)
    chart_sheet = workbook[graph_sheet_name]

    img = Image(image_file)
    img.anchor = "A1"
    chart_sheet.add_image(img)

    workbook.save(file_name)

def export_to_excel(df, file_name):
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Amortization Schedule")
    workbook = load_workbook(file_name)
    sheet = workbook["Amortization Schedule"]

    dollar_columns = ["Monthly Payment", "Principal Paid", "Interest Paid", "Total Interest Paid", "Remaining Balance"]
    for col_name in dollar_columns:
        if col_name in df.columns:
            col_letter = sheet.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    workbook.save(file_name)

if __name__ == "__main__":
    while True:
        try:
            loan_amount = float(input("Enter the total loan amount: "))
            interest_rate = float(input("Enter the annual interest rate (in %): "))
            loan_term = int(input("Enter the loan term (in years): "))
            down_payment = float(input("Enter the down payment amount (optional, default is 0): ") or 0)
            trade_in_value = float(input("Enter the trade-in value (optional, default is 0): ") or 0)
            extra_payment = float(input("Enter the extra monthly payment (optional, default is 0): ") or 0)
            file_name = input("Enter the base name for the output files (e.g., 'auto_loan'): ")
            break
        except Exception as e:
            print(f"Error: {e}. Please try again.")

    image_file = f"{file_name}.png"
    excel_file = f"{file_name}.xlsx"

    df = auto_loan_calculator(loan_amount, interest_rate, loan_term, down_payment, trade_in_value, extra_payment)
    plot_loan_amortization(df, image_file)
    export_to_excel(df, excel_file)
    embed_chart_in_excel(excel_file, image_file)
    auto_adjust_column_width(excel_file)

    print(f"Auto loan details saved to {excel_file} with an amortization graph embedded.")
