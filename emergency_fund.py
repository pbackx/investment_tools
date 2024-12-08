import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def calculate_emergency_fund(monthly_expenses, coverage_months, current_savings=0, contribution_amount=0, contribution_frequency="monthly"):
    # Calculate total target emergency fund
    target_fund = monthly_expenses * coverage_months

    # Map contribution frequency to periods
    freq_map = {"daily": 365, "weekly": 52, "bi-weekly": 26, "monthly": 12}
    contribution_periods_per_year = freq_map.get(contribution_frequency.lower(), 12)
    contribution_per_period = contribution_amount / (12 / contribution_periods_per_year)

    # Savings progress
    balance = current_savings
    months_needed = 0
    results = []

    while balance < target_fund:
        months_needed += 1
        balance += contribution_per_period * (12 / contribution_periods_per_year)

        results.append({
            "Month": months_needed,
            "Savings Balance": balance,
            "Target Fund": target_fund,
            "Remaining Amount": max(0, target_fund - balance)
        })

    # Create DataFrame
    df = pd.DataFrame(results)
    return df, target_fund

def plot_emergency_fund(df, file_name):
    # Plot savings progress
    plt.figure(figsize=(12, 7))
    plt.plot(df["Month"], df["Savings Balance"], label="Savings Balance", color="green")
    plt.axhline(y=df["Target Fund"].iloc[0], color="blue", linestyle="--", label="Target Fund")
    plt.title("Emergency Fund Savings Progress")
    plt.xlabel("Month")
    plt.ylabel("Amount ($)")
    plt.legend(loc="upper left")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)
    plt.close()

def auto_adjust_column_width(file_name):
    workbook = load_workbook(file_name)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    workbook.save(file_name)

def embed_chart_in_excel(file_name, image_file):
    workbook = load_workbook(file_name)
    graph_sheet_name = "Graph"
    if graph_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(graph_sheet_name)
    chart_sheet = workbook[graph_sheet_name]

    img = Image(image_file)
    img.anchor = "A1"
    chart_sheet.add_image(img)

    workbook.save(file_name)

def export_to_excel(df, file_name):
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Savings Progress")
    workbook = load_workbook(file_name)
    sheet = workbook["Savings Progress"]

    # Apply dollar formatting
    dollar_columns = ["Savings Balance", "Target Fund", "Remaining Amount"]
    for col_name in dollar_columns:
        if col_name in df.columns:
            col_letter = sheet.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    workbook.save(file_name)

if __name__ == "__main__":
    while True:
        try:
            monthly_expenses = float(input("Enter your total monthly expenses: "))
            coverage_months = int(input("Enter the desired coverage period (in months): "))
            current_savings = float(input("Enter your current savings amount (optional, default is 0): ") or 0)
            contribution_amount = float(input("Enter your planned contribution amount (optional, default is 0): ") or 0)
            contribution_frequency = input("Enter the contribution frequency (daily, weekly, bi-weekly, monthly): ").lower()
            file_name = input("Enter the base name for the output files (e.g., 'emergency_fund'): ")
            break
        except Exception as e:
            print(f"Error: {e}. Please try again.")

    image_file = f"{file_name}.png"
    excel_file = f"{file_name}.xlsx"

    df, target_fund = calculate_emergency_fund(monthly_expenses, coverage_months, current_savings, contribution_amount, contribution_frequency)
    plot_emergency_fund(df, image_file)
    export_to_excel(df, excel_file)
    embed_chart_in_excel(excel_file, image_file)
    auto_adjust_column_width(excel_file)

    print(f"Emergency fund savings details saved to {excel_file} with a progress graph embedded.")
    print(f"Target Emergency Fund: ${target_fund:,.2f}")
