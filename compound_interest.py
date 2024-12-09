import math
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def compound_interest(principal, annual_rate, contribution, frequency, total_duration, is_duration_in_years, annual_increase=0, inflation_rate=0):
    freq_map = {"daily": 365, "weekly": 52, "bi-weekly": 26, "monthly": 12, "yearly": 1}
    periods_per_year = freq_map.get(frequency, 12)
    total_periods = total_duration * periods_per_year if is_duration_in_years else total_duration * periods_per_year // 12

    periodic_rate = (annual_rate / 100) / periods_per_year
    results = []
    balance = principal
    total_contributions = principal
    total_interest_earned = 0

    for period in range(1, total_periods + 1):
        # Add contributions at the start of each period
        balance += contribution
        total_contributions += contribution

        # Calculate interest for the period
        interest = balance * periodic_rate
        balance += interest
        total_interest_earned += interest

        # Determine the current month
        current_month = (period - 1) * 12 // periods_per_year + 1

        # Apply inflation adjustment if applicable
        real_balance = balance / ((1 + inflation_rate / 100) ** (current_month / 12)) if inflation_rate > 0 else None

        # Append results for the period
        result = {
            "Period": period,
            "Month": current_month,
            "Year": (current_month - 1) // 12 + 1,
            "Principal Paid": total_contributions,
            "Interest Paid (This Period)": interest,
            "Total Interest Paid": total_interest_earned,
            "Balance": balance,
        }
        if inflation_rate > 0:
            result["Real Balance"] = real_balance

        results.append(result)

        # Apply annual contribution increase
        if period % periods_per_year == 0:
            contribution *= (1 + annual_increase / 100)

    return pd.DataFrame(results)

def plot_investment_growth(df, file_name, display_by, inflation_rate):
    plt.figure(figsize=(12, 7))
    x_label = "Year" if display_by == "years" else "Period"
    plt.plot(df[x_label], df["Balance"], label="Nominal Balance", color="blue")
    if inflation_rate > 0 and "Real Balance" in df.columns:
        plt.plot(df[x_label], df["Real Balance"], label="Real Balance (Inflation Adjusted)", linestyle="--", color="orange")
    plt.title("Investment Growth Over Time")
    plt.xlabel(x_label)
    plt.ylabel("Balance ($)")
    plt.legend(loc="upper left")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(file_name)
    plt.close()

def auto_adjust_column_width(file_name):
    workbook = load_workbook(file_name)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    workbook.save(file_name)

def embed_chart_in_excel(file_name, image_file):
    workbook = load_workbook(file_name)
    graph_sheet_name = "Graph"
    if graph_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(graph_sheet_name)
    chart_sheet = workbook[graph_sheet_name]

    img = Image(image_file)
    img.anchor = "A1"
    chart_sheet.add_image(img)

    workbook.save(file_name)

def export_to_excel(df, file_name):
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Detailed Data")
    workbook = load_workbook(file_name)
    sheet = workbook["Detailed Data"]

    dollar_columns = ["Principal Paid", "Interest Paid (This Period)", "Total Interest Paid", "Balance"]
    if "Real Balance" in df.columns:
        dollar_columns.append("Real Balance")
        
    for col_name in dollar_columns:
        if col_name in df.columns:
            col_letter = sheet.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
            for row in range(2, sheet.max_row + 1):
                cell = sheet[f"{col_letter}{row}"]
                cell.number_format = '"$"#,##0.00'

    workbook.save(file_name)

if __name__ == "__main__":
    while True:
        try:
            principal = float(input("Enter the initial principal amount: "))
            annual_rate = float(input("Enter the annual interest rate (in %): "))
            contribution = float(input("Enter the contribution amount per period: "))
            frequency = input("Enter the contribution frequency (daily, weekly, bi-weekly, monthly, yearly): ").lower()
            duration_input = input("Enter the total duration (e.g., '12 months' or '5 years'): ").lower()
            duration_parts = duration_input.split()
            duration = int(duration_parts[0])
            is_duration_in_years = "year" in duration_parts[1]
            display_by = input("Display results by 'months' or 'years': ").lower()
            annual_increase = float(input("Enter the annual contribution increase rate (in %, default is 0): ") or 0)
            inflation_rate = float(input("Enter the inflation rate (in %, default is 0): ") or 0)
            file_name = input("Enter the base name for the output files (e.g., 'results'): ")
            break
        except Exception as e:
            print(f"Error: {e}. Please try again.")

    image_file = f"{file_name}.png"
    excel_file = f"{file_name}.xlsx"

    df = compound_interest(principal, annual_rate, contribution, frequency, duration, is_duration_in_years, annual_increase, inflation_rate)
    plot_investment_growth(df, image_file, display_by, inflation_rate)
    export_to_excel(df, excel_file)
    embed_chart_in_excel(excel_file, image_file)
    auto_adjust_column_width(excel_file)
